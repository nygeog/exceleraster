from PIL import Image
import numpy as np
import pandas as pd
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl import Workbook
from openpyxl.styles import Color, PatternFill, Font, Border
from openpyxl.styles import colors
from openpyxl.cell import Cell
import matplotlib 


def rgb2hex(r, g, b):
    return '#{:02x}{:02x}{:02x}'.format(r, g, b)

def load_image( infilename ) :
    img = Image.open( infilename )
    img.load()
    data = np.asarray(img, dtype="int32" )
    return data

def save_image( npdata, outfilename ) :
    img = Image.fromarray(np.asarray(np.clip(npdata,0,255), dtype="uint8"), "L" )
    img.save( outfilename )


def tohex(array):
    array = np.asarray(array, dtype='uint32')
    return ((array[:, :, 0]<<16) + (array[:, :, 1]<<8) + array[:, :, 2])


def rgb2hex(i):
    i = i.tolist()
    r, g, b = i#[0], i[1], 1[2]
    return '#{:02x}{:02x}{:02x}'.format(r, g, b)



get_hex_color = lambda t: rgb2hex(t[0], t[1], t[2])
vfunc = np.vectorize(get_hex_color)

data = load_image('image.jpg')
#data = load_image('park.jpg')

#for i in [image.jpg']
df = pd.DataFrame.from_records(data)

dfhex = df.applymap(rgb2hex)

rows = dataframe_to_rows(dfhex)

wb = Workbook()
ws = wb.active

coord_list = []

for r_idx, row in enumerate(rows, 1):
    for c_idx, value in enumerate(row, 1):
        try:
            color = value[1:7]
            cell = ws.cell(row=r_idx, column=c_idx, value=None)
            coord = ''.join([cell.column, str(cell.row)])
            coord_list.append(coord)
            #color = value[1:7]
            ws[coord].fill=PatternFill(start_color=color, end_color=color, fill_type="solid")
            ws.column_dimensions[cell.column].width = 2.1 # fix this so only iterates at col, rather than at every cel
        except Exception: 
              pass

wb.save('test.xlsx')
